from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

wb = Workbook()
ws = wb.active
ws.title = "品鉴记录"
ls = wb.create_sheet("_Lists")
ls.sheet_state = 'hidden'

LISTS = {
    'process':    ['水洗','日晒','蜜处理','厌氧'],
    'roast':      ['浅','中浅','中','中深','深'],
    'brew':       ['手冲','意式','冷萃','杯测','其他'],
    'color':      ['浅金黄','琥珀','深棕','黑褐'],
    'clarity':    ['清澈','略浑浊','浑浊'],
    'style':      ['花香果香型','醇厚甜感型','发酵复杂型','坚果巧克力型'],
    'aroma_imp':  ['轻盈花香','浓郁果香','发酵感强','坚果甜香','烟熏感','内敛沉稳'],
    'ar_citrus':  ['柠檬','葡萄柚','橙子','青柠','苹果','梨'],
    'ar_berry':   ['草莓','覆盆子','蓝莓','桃子','芒果','百香果','菠萝'],
    'ar_floral':  ['茉莉','玫瑰','橙花','红茶','绿茶','洋甘菊'],
    'ar_sweet':   ['焦糖','红糖','蜂蜜','巧克力','可可','榛子','杏仁','花生'],
    'ar_ferment': ['酒香','醋感','发酵浆果','烟草','雪松','泥土','香料'],
    'acid_lvl':   ['低','中低','中','高','强'],
    'acid_type':  ['柑橘酸','苹果酸','莓果酸','发酵酸'],
    'body':       ['轻盈','中等','厚重'],
    'sweet':      ['弱','中','强'],
    'texture':    ['水感','顺滑','丝绒感','奶油感','糖浆感','涩感','干燥感'],
    'finish':     ['短','略短','略长','长'],
    'rating':     ['★★★','★★','★','—'],
}

LIST_REFS = {}
for ci, (key, vals) in enumerate(LISTS.items(), 1):
    cl = get_column_letter(ci)
    for ri, v in enumerate(vals, 1):
        ls.cell(row=ri, column=ci, value=v)
    LIST_REFS[key] = f'_Lists!${cl}$1:${cl}${len(vals)}'

C_SEC   = "3B2314"
C_SUB   = "F5EFE8"
C_LABEL = "F2F0ED"
C_INPUT = "FFFFFF"
C_EMPTY = "F5F5F2"

def F(h): return PatternFill("solid", fgColor=h)
TH = Side(style='thin', color='DDDDD8')
def BRD(): return Border(left=TH, right=TH, top=TH, bottom=TH)

def SC(cell, val=None, bold=False, sz=10, col="1A1A1A",
       bg=C_INPUT, ha='left', va='center', wrap=False, ind=0, nf=None):
    if val is not None: cell.value = val
    cell.font = Font(bold=bold, size=sz, color=col)
    cell.fill = F(bg)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap, indent=ind)
    cell.border = BRD()
    if nf: cell.number_format = nf

def section(row, text):
    ws.row_dimensions[row].height = 24
    SC(ws.cell(row, 2), text, bold=True, sz=11, col="FFFFFF", bg=C_SEC, ind=1)
    for c in range(3, 13): SC(ws.cell(row, c), bg=C_SEC)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)

def subsec(row, text):
    ws.row_dimensions[row].height = 18
    SC(ws.cell(row, 2), text, sz=9, col=C_SEC, bg=C_SUB, ind=1)
    for c in range(3, 13): SC(ws.cell(row, c), bg=C_SUB)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)

def li(row, label, dd=None, height=22, is_date=False):
    ws.row_dimensions[row].height = height
    SC(ws.cell(row, 2), label, sz=10, col="333333", bg=C_LABEL, ind=1)
    ic = ws.cell(row, 3)
    SC(ic, bg=C_INPUT, ind=1, nf=('YYYY-MM-DD' if is_date else None))
    for c in range(4, 13): SC(ws.cell(row, c), bg=C_INPUT)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
    if dd:
        dv = DataValidation(type="list", formula1=LIST_REFS[dd], allow_blank=True)
        ws.add_data_validation(dv); dv.add(ic)
    return ic

def multi(row, label, dd, n, height=22):
    ws.row_dimensions[row].height = height
    SC(ws.cell(row, 2), label, sz=10, col="333333", bg=C_LABEL, ind=1)
    dv = DataValidation(type="list", formula1=LIST_REFS[dd], allow_blank=True)
    ws.add_data_validation(dv)
    slots = min(n, 10)
    for ci in range(slots):
        c = ws.cell(row, 3 + ci)
        SC(c, bg=C_INPUT, ha='center'); dv.add(c)
    rs = 3 + slots
    if rs <= 12:
        for c in range(rs, 13): SC(ws.cell(row, c), bg=C_EMPTY)
        if rs < 12:
            ws.merge_cells(start_row=row, start_column=rs, end_row=row, end_column=12)

def lbl(row, text):
    ws.row_dimensions[row].height = 17
    SC(ws.cell(row, 2), text, sz=9, col="666666", bg=C_LABEL, ind=1)
    for c in range(3, 13): SC(ws.cell(row, c), bg=C_LABEL)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)

def ta(row, height=52):
    ws.row_dimensions[row].height = height
    c = ws.cell(row, 2)
    SC(c, bg=C_INPUT, va='top', wrap=True, ind=1)
    for cc in range(3, 13): SC(ws.cell(row, cc), bg=C_INPUT)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)

def sp(row, h=8):
    ws.row_dimensions[row].height = h
    for c in range(1, 13):
        ws.cell(row, c).fill = F('FFFFFF')

ws.column_dimensions['A'].width = 0.6
ws.column_dimensions['B'].width = 15
for c in ['C','D','E','F','G','H','I','J','K','L']:
    ws.column_dimensions[c].width = 8.8
ws.sheet_view.showGridLines = False

r = 1
ws.row_dimensions[r].height = 36
ws.merge_cells(f'A{r}:L{r}')
tc = ws[f'A{r}']
tc.value = '咖啡品鉴记录'
tc.font = Font(bold=True, size=16, color=C_SEC)
tc.alignment = Alignment(horizontal='center', vertical='center')
tc.fill = F('FFFFFF')

r+=1; sp(r)

# 基本信息
r+=1; section(r, '基本信息')
r+=1; li(r, '国家')
r+=1; li(r, '产区')
r+=1; li(r, '庄园 / 合作社')
r+=1; li(r, '品种（Variety）')
r+=1; li(r, '处理法', dd='process')
r+=1; li(r, '具体方式')
r+=1; li(r, '烘焙度', dd='roast')
r+=1; li(r, '烘焙商')
r+=1; li(r, '品鉴场所')
r+=1; li(r, '冲煮方式', dd='brew')
r+=1; li(r, '品鉴日期', is_date=True)
r+=1; sp(r)

# 外观
r+=1; section(r, '外观')
r+=1; li(r, '颜色', dd='color')
r+=1; li(r, '透明度', dd='clarity')
r+=1; sp(r)

# 香气
r+=1; section(r, '香气')
r+=1; li(r, '干香印象', dd='aroma_imp')
r+=1; subsec(r, '香气特征（多选）— 每个下拉格选一个词')
r+=1; multi(r, '柑橘·果酸', 'ar_citrus', 6)
r+=1; multi(r, '浆果·热带', 'ar_berry', 7)
r+=1; multi(r, '花香·茶', 'ar_floral', 6)
r+=1; multi(r, '甜香·坚果', 'ar_sweet', 8)
r+=1; multi(r, '发酵·其他', 'ar_ferment', 7)
r+=1; lbl(r, '自由描述')
r+=1; ta(r, 56)
r+=1; sp(r)

# 味道
r+=1; section(r, '味道')
r+=1; li(r, '酸质强度', dd='acid_lvl')
r+=1; li(r, '酸质类型', dd='acid_type')
r+=1; li(r, '醇厚度（Body）', dd='body')
r+=1; li(r, '甜感', dd='sweet')
r+=1; multi(r, '口感质地', 'texture', 7)
r+=1; li(r, '余韵长度', dd='finish')
r+=1; lbl(r, '余韵描述')
r+=1; ta(r, 44)
r+=1; sp(r)

# 综合判断
r+=1; section(r, '综合判断')
r+=1; li(r, '风格倾向', dd='style')
r+=1; li(r, '评分', dd='rating')
r+=1; sp(r)

# 备注
r+=1; section(r, '备注')
r+=1; li(r, '品鉴场所背景')
r+=1; lbl(r, '场景故事')
r+=1; ta(r, 52)
r+=1; lbl(r, '与其他咖啡的对比')
r+=1; ta(r, 44)
r+=1; lbl(r, '一句话总结')
r+=1; ta(r, 44)

ws.page_setup.orientation = 'portrait'
ws.page_setup.paperSize = 9
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.6
ws.page_margins.bottom = 0.6
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.print_area = f'A1:L{r}'

# 词汇参考 sheet
ref = wb.create_sheet("词汇参考")
ref.sheet_view.showGridLines = False
ref.column_dimensions['A'].width = 1.5
ref.column_dimensions['B'].width = 16
ref.column_dimensions['C'].width = 20
ref.column_dimensions['D'].width = 26

REF_DATA = [
    ('处理法', [
        ('水洗 Washed', '发酵后水洗去果肉', '干净、酸质清晰、花香突出'),
        ('日晒 Natural', '带果肉直接晒干', '果香浓郁、发酵感、甜感强'),
        ('蜜处理 Honey', '保留部分果肉干燥', '介于水洗和日晒之间'),
        ('厌氧 Anaerobic', '密封无氧环境发酵', '发酵感极强，风格特殊'),
    ]),
    ('烘焙度', [
        ('浅烘', 'Light Roast', '果酸明亮，花香保留，适合手冲'),
        ('中浅烘', 'Medium-Light', '酸甜平衡，香气丰富'),
        ('中烘', 'Medium Roast', '平衡，焦糖感出现'),
        ('中深烘', 'Medium-Dark', '苦感增加，坚果巧克力感'),
        ('深烘', 'Dark Roast', '苦味主导，烟熏感，适合意式'),
    ]),
    ('酸质类型', [
        ('柑橘酸', '明亮、尖锐', '常见于水洗埃塞'),
        ('苹果酸', '清爽、圆润', '常见于高海拔产区'),
        ('莓果酸', '果味感强', '常见于日晒豆'),
        ('发酵酸', '复杂、有层次', '常见于厌氧处理'),
    ]),
    ('香气·柑橘果酸', [(v,'','') for v in LISTS['ar_citrus']]),
    ('香气·浆果热带', [(v,'','') for v in LISTS['ar_berry']]),
    ('香气·花香茶', [(v,'','') for v in LISTS['ar_floral']]),
    ('香气·甜香坚果', [(v,'','') for v in LISTS['ar_sweet']]),
    ('香气·发酵其他', [(v,'','') for v in LISTS['ar_ferment']]),
    ('醇厚度', [(v,'','') for v in LISTS['body']]),
    ('口感质地', [(v,'','') for v in LISTS['texture']]),
    ('余韵长度', [(v,'','') for v in LISTS['finish']]),
]

rr = 1
ref.row_dimensions[rr].height = 32
ref.merge_cells(f'A{rr}:D{rr}')
h = ref.cell(rr, 1, '咖啡品鉴词汇参考')
h.font = Font(bold=True, size=14, color=C_SEC)
h.alignment = Alignment(horizontal='center', vertical='center')
h.fill = F('FFFFFF')
rr += 2

TH2 = Side(style='thin', color='DDDDD8')
def BRD2(): return Border(left=TH2, right=TH2, top=TH2, bottom=TH2)

def rc(cell, val, bg, bold=False, sz=10, col='1A1A1A', italic=False):
    cell.value = val
    cell.font = Font(bold=bold, size=sz, color=col, italic=italic)
    cell.fill = F(bg)
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    cell.border = BRD2()

for cat, items in REF_DATA:
    ref.row_dimensions[rr].height = 20
    ref.merge_cells(f'A{rr}:D{rr}')
    rc(ref.cell(rr, 1), cat, C_SEC, bold=True, sz=10, col='FFFFFF')
    rr += 1
    for i, (zh, jp, desc) in enumerate(items):
        ref.row_dimensions[rr].height = 17
        bg = 'FAFAFA' if i % 2 == 0 else 'FFFFFF'
        rc(ref.cell(rr, 2), zh, bg, sz=10)
        if desc:
            rc(ref.cell(rr, 3), jp, bg, sz=9, col='888888', italic=True)
            rc(ref.cell(rr, 4), desc, bg, sz=9, col='555555')
        else:
            ref.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
        rr += 1
    rr += 1

out = '/home/user/alcoholic/coffee/templates/coffee-tasting-form.xlsx'
wb.save(out)
print(f"Saved: {out}  (form rows: {r})")
