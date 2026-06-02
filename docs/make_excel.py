from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "品评记录"
ls = wb.create_sheet("_Lists")
ls.sheet_state = 'hidden'

LISTS = {
    'grade':    ['纯米大吟酿','大吟酿','纯米吟酿','吟酿','特别纯米','纯米','特别本酿造','本酿造'],
    'yeast':    ['速醸','生酛','山廃','不明'],
    'features': ['生酒','原酒','无过滤','浊酒','古酒'],
    'temp':     ['冷酒','常温','燗酒'],
    'clarity':  ['晶莹透澈','略有浑浊','明显浑浊'],
    'color':    ['水晶白','银白','金黄','淡黄','黄玉色','橙色','棕褐'],
    'aroma_imp':['年轻清新','爽快清爽','华丽','丰盈','芳醇','沉稳','有熟成感'],
    'ar_fruit': ['葡萄柚','黄苹果','西洋梨','白桃','香蕉','哈密瓜','麝香葡萄','荔枝'],
    'ar_floral':['忍冬花','紫罗兰','金合欢','椴树花','细叶芹','青竹','新绿'],
    'ar_grain': ['蒸米饭','新打麻糍','米粉','汤圆','杏仁豆腐','生奶油','酸奶油','酸奶','发酵黄油','奶油芝士'],
    'ar_wood':  ['桧木','香木','月桂叶','丁香','肉桂'],
    'ar_aged':  ['红茶','焦糖','咖啡','酱油','蜂蜜','黑巧克力','栗子','杏仁','黏土','石灰','碘香'],
    'body':     ['轻盈','略轻','略重','厚重'],
    'sweet':    ['淡甜','优雅','圆润','丰盈','强劲'],
    'acid':     ['爽快','柔和','圆润','锐利','强劲'],
    'bitter':   ['轻微','增醇厚感','带鲜味','明显'],
    'balance':  ['顺滑','活泼','干爽','圆润','黏稠','丰润','有厚度','强劲'],
    'finish':   ['短','略短','略长','长'],
    'ssi':      ['薫酒（高香·淡味）','爽酒（低香·淡味）','醇酒（低香·浓味）','熟酒（高香·浓味）'],
    'rating':   ['★★★','★★','★','—'],
}

LIST_REFS = {}
for ci, (key, vals) in enumerate(LISTS.items(), 1):
    cl = get_column_letter(ci)
    for ri, v in enumerate(vals, 1):
        ls.cell(row=ri, column=ci, value=v)
    LIST_REFS[key] = f'_Lists!${cl}$1:${cl}${len(vals)}'

# Colors
C_SEC   = "1A5C40"
C_SUB   = "E8F5EE"
C_LABEL = "F2F2ED"
C_INPUT = "FFFFFF"
C_EMPTY = "F5F5F2"

def F(h): return PatternFill("solid", fgColor=h)
TH = Side(style='thin', color='DDDDD8')
NO = Side(style=None)
def BRD(): return Border(left=TH, right=TH, top=TH, bottom=TH)

def SC(cell, val=None, bold=False, sz=10, col="1A1A1A",
       bg=C_INPUT, ha='left', va='center', wrap=False, ind=0, nf=None):
    if val is not None: cell.value = val
    cell.font = Font(bold=bold, size=sz, color=col)
    cell.fill = F(bg)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap, indent=ind)
    cell.border = BRD()
    if nf: cell.number_format = nf

def section(row, text):
    ws.row_dimensions[row].height = 24
    SC(ws.cell(row, 2), text, bold=True, sz=11, col="FFFFFF", bg=C_SEC, ind=1)
    for c in range(3, 13): SC(ws.cell(row, c), bg=C_SEC)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)

def subsec(row, text):
    ws.row_dimensions[row].height = 18
    SC(ws.cell(row, 2), text, sz=9, col="1A5C40", bg=C_SUB, ind=1)
    for c in range(3, 13): SC(ws.cell(row, c), bg=C_SUB)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)

def li(row, label, dd=None, height=22, is_date=False):
    ws.row_dimensions[row].height = height
    SC(ws.cell(row, 2), label, sz=10, col="333333", bg=C_LABEL, ind=1)
    ic = ws.cell(row, 3)
    SC(ic, bg=C_INPUT, ind=1, nf=('YYYY-MM-DD' if is_date else None))
    for c in range(4, 13): SC(ws.cell(row, c), bg=C_INPUT)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
    if dd:
        dv = DataValidation(type="list", formula1=LIST_REFS[dd], allow_blank=True)
        ws.add_data_validation(dv); dv.add(ic)
    return ic

def pct(row, label):
    ws.row_dimensions[row].height = 22
    SC(ws.cell(row, 2), label, sz=10, col="333333", bg=C_LABEL, ind=1)
    ic = ws.cell(row, 3)
    SC(ic, bg=C_INPUT, ha='right', ind=1)
    for c in range(4, 12): SC(ws.cell(row, c), bg=C_INPUT)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=11)
    SC(ws.cell(row, 12), '%', sz=9, col="AAAAAA", bg=C_LABEL, ha='left', ind=1)

def multi(row, label, dd, n, height=22):
    ws.row_dimensions[row].height = height
    SC(ws.cell(row, 2), label, sz=10, col="333333", bg=C_LABEL, ind=1)
    dv = DataValidation(type="list", formula1=LIST_REFS[dd], allow_blank=True)
    ws.add_data_validation(dv)
    slots = min(n, 10)
    for ci in range(slots):
        c = ws.cell(row, 3 + ci)
        SC(c, bg=C_INPUT, ha='center'); dv.add(c)
    rs = 3 + slots
    if rs <= 12:
        for c in range(rs, 13): SC(ws.cell(row, c), bg=C_EMPTY)
        if rs < 12:
            ws.merge_cells(start_row=row, start_column=rs, end_row=row, end_column=12)

def lbl(row, text):
    ws.row_dimensions[row].height = 17
    SC(ws.cell(row, 2), text, sz=9, col="666666", bg=C_LABEL, ind=1)
    for c in range(3, 13): SC(ws.cell(row, c), bg=C_LABEL)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)

def ta(row, height=52):
    ws.row_dimensions[row].height = height
    c = ws.cell(row, 2)
    SC(c, bg=C_INPUT, va='top', wrap=True, ind=1)
    for cc in range(3, 13): SC(ws.cell(row, cc), bg=C_INPUT)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)

def sp(row, h=8):
    ws.row_dimensions[row].height = h
    for c in range(1, 13):
        ws.cell(row, c).fill = F('FFFFFF')

# Column widths
ws.column_dimensions['A'].width = 0.6
ws.column_dimensions['B'].width = 15
for c in ['C','D','E','F','G','H','I','J','K','L']:
    ws.column_dimensions[c].width = 8.8
ws.sheet_view.showGridLines = False

# Title
r = 1
ws.row_dimensions[r].height = 36
ws.merge_cells(f'A{r}:L{r}')
tc = ws[f'A{r}']
tc.value = '清酒品评记录'
tc.font = Font(bold=True, size=16, color=C_SEC)
tc.alignment = Alignment(horizontal='center', vertical='center')
tc.fill = F('FFFFFF')

r+=1; sp(r)

# 基本信息
r+=1; section(r, '基本信息')
r+=1; li(r, '酒款名')
r+=1; li(r, '蔵元')
r+=1; li(r, '产地')
r+=1; li(r, '特定名称', dd='grade')
r+=1; pct(r, '精米步合')
r+=1; pct(r, '酒精度')
r+=1; li(r, '使用酒米')
r+=1; li(r, '酒母', dd='yeast')
r+=1; multi(r, '其他特征', 'features', 5)
r+=1; li(r, '品饮日期', is_date=True)
r+=1; li(r, '品饮场所')
r+=1; li(r, '品饮温度', dd='temp')
r+=1; sp(r)

# 外观
r+=1; section(r, '外观')
r+=1; li(r, '清澄度', dd='clarity')
r+=1; li(r, '色调', dd='color')
r+=1; lbl(r, '备注')
r+=1; ta(r, 40)
r+=1; sp(r)

# 香气
r+=1; section(r, '香气')
r+=1; li(r, '第一印象', dd='aroma_imp')
r+=1; subsec(r, '香气特征（选 5～11 项）— 每个下拉格选一个词')
r+=1; multi(r, '果実系', 'ar_fruit', 8)
r+=1; multi(r, '花·植物系', 'ar_floral', 7)
r+=1; multi(r, '米·穀物·乳系', 'ar_grain', 10)
r+=1; multi(r, '木·辛香系', 'ar_wood', 5)
r+=1; multi(r, '熟成·その他', 'ar_aged', 10)
r+=1; lbl(r, '自己的描述（香气）')
r+=1; ta(r, 56)
r+=1; sp(r)

# 味道
r+=1; section(r, '味道')
r+=1; li(r, '重量感', dd='body')
r+=1; li(r, '甜感（甘み）', dd='sweet')
r+=1; li(r, '酸感（酸み）', dd='acid')
r+=1; li(r, '苦感（苦み）', dd='bitter')
r+=1; multi(r, '平衡感（バランス）', 'balance', 8)
r+=1; li(r, '余韵长度', dd='finish')
r+=1; lbl(r, '余韵特征')
r+=1; ta(r, 44)
r+=1; sp(r)

# 综合判断
r+=1; section(r, '综合判断')
r+=1; li(r, 'SSI 风格类型', dd='ssi')
r+=1; lbl(r, '特定名称推断（仅凭口感）')
r+=1; ta(r, 44)
r+=1; li(r, '评分', dd='rating')
r+=1; sp(r)

# 备注
r+=1; section(r, '备注')
r+=1; li(r, '搭配料理')
r+=1; lbl(r, '背景 / 故事')
r+=1; ta(r, 60)
r+=1; lbl(r, '与其他酒的对比')
r+=1; ta(r, 44)
r+=1; lbl(r, '一句话总结')
r+=1; ta(r, 44)

# Page setup
ws.page_setup.orientation = 'portrait'
ws.page_setup.paperSize = 9
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.6
ws.page_margins.bottom = 0.6
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.print_area = f'A1:L{r}'

# 词汇参考 sheet
ref = wb.create_sheet("词汇参考")
ref.sheet_view.showGridLines = False
ref.column_dimensions['A'].width = 1.5
ref.column_dimensions['B'].width = 16
ref.column_dimensions['C'].width = 16
ref.column_dimensions['D'].width = 26

REF_DATA = [
    ('法定分类（特定名称）', [
        ('纯米大吟酿', '純米大吟醸', '精米步合 ≤50%，全米，最高等级'),
        ('大吟酿',    '大吟醸',    '精米步合 ≤50%，含少量酿造酒精'),
        ('纯米吟酿',  '純米吟醸',  '精米步合 ≤60%，全米，吟醸香'),
        ('吟酿',      '吟醸',      '精米步合 ≤60%，含少量酿造酒精'),
        ('特别纯米',  '特別純米',  '精米步合 ≤60% 或特殊制法，全米'),
        ('纯米',      '純米',      '无精米步合规定，全米酿造'),
        ('特别本酿造','特別本醸造','精米步合 ≤60%，含少量酿造酒精'),
        ('本酿造',    '本醸造',    '精米步合 ≤70%，含少量酿造酒精'),
    ]),
    ('SSI 风格四类型', [
        ('薫酒 くんしゅ',   '高香·淡味', '大吟酿系，果香华丽，冷饮最佳'),
        ('爽酒 そうしゅ',   '低香·淡味', '本酿造·普通酒，清爽干净'),
        ('醇酒 じゅんしゅ', '低香·浓味', '生酛·山廃，旨味饱满，适燗酒'),
        ('熟酒 じゅくしゅ', '高香·浓味', '古酒·熟成系，琥珀色，复杂'),
    ]),
    ('香气第一印象', [(v,'','') for v in LISTS['aroma_imp']]),
    ('果実系香气',   [(v,'','') for v in LISTS['ar_fruit']]),
    ('花·植物系香气',[(v,'','') for v in LISTS['ar_floral']]),
    ('米·穀物·乳系香气',[(v,'','') for v in LISTS['ar_grain']]),
    ('木·辛香系香气',[(v,'','') for v in LISTS['ar_wood']]),
    ('熟成·その他香气',[(v,'','') for v in LISTS['ar_aged']]),
    ('重量感', [(v,'','') for v in LISTS['body']]),
    ('甜感（甘み）',[(v,'','') for v in LISTS['sweet']]),
    ('酸感（酸み）',[(v,'','') for v in LISTS['acid']]),
    ('苦感（苦み）',[(v,'','') for v in LISTS['bitter']]),
    ('平衡感（バランス）',[(v,'','') for v in LISTS['balance']]),
    ('余韵长度', [(v,'','') for v in LISTS['finish']]),
]

rr = 1
ref.row_dimensions[rr].height = 32
ref.merge_cells(f'A{rr}:D{rr}')
h = ref.cell(rr, 1, '品评词汇参考')
h.font = Font(bold=True, size=14, color=C_SEC)
h.alignment = Alignment(horizontal='center', vertical='center')
h.fill = F('FFFFFF')
rr += 2

def rc(cell, val, bg, bold=False, sz=10, col='1A1A1A', italic=False):
    cell.value = val
    cell.font = Font(bold=bold, size=sz, color=col, italic=italic)
    cell.fill = F(bg)
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    cell.border = BRD()

for cat, items in REF_DATA:
    ref.row_dimensions[rr].height = 20
    ref.merge_cells(f'A{rr}:D{rr}')
    rc(ref.cell(rr, 1), cat, C_SEC, bold=True, sz=10, col='FFFFFF')
    rr += 1
    for i, (zh, jp, desc) in enumerate(items):
        ref.row_dimensions[rr].height = 17
        bg = 'FAFAFA' if i % 2 == 0 else 'FFFFFF'
        rc(ref.cell(rr, 2), zh, bg, sz=10)
        if desc:
            rc(ref.cell(rr, 3), jp, bg, sz=9, col='888888', italic=True)
            rc(ref.cell(rr, 4), desc, bg, sz=9, col='555555')
        else:
            ref.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
        rr += 1
    rr += 1

out = r'C:\Users\qianx\Documents\GitHub\alcoholic\docs\sake-tasting-form.xlsx'
wb.save(out)
print(f"Saved: {out}  (form rows: {r})")